import openpyxl

class GetDate:
    def xlsx(self):
        wk = openpyxl.load_workbook('testcase.xlsx')
        sheet1 = wk['Sheet1']
        # vs = sheet1.cell(1, 1).value
        # print(vs)
        datelist=[]
        for x in range(2,sheet1.max_row+1):
            date1={}
            for y in range(1,sheet1.max_column+1):
                vs=date1[sheet1.cell(1,y).value]=sheet1.cell(x,y).value
                print(vs)
                datelist.append(date1)
        print(datelist)
        return datelist


# if __name__ == '__main__':
#     xlsx()


# wk=openpyxl.load_workbook('testcase.xlsx')
# sheet1=wk['Sheet1']
#
# datelist=[]
# for x in range(2,sheet1.max_row+1):
#     date={}
#     for y in range(1,sheet1.max_column+1):
#         date[sheet1.cell(1,y).value] = sheet1.cell(x,y).value
#         datelist.append(date)





















